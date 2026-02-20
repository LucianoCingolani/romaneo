from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from app import models
from app.forms import TaskForm
import csv
from django.utils import timezone
from django.template.loader import get_template
from xhtml2pdf import pisa
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Create your views here.

def home(request):
    lista_procesos = models.Procesos.objects.all()

    if request.method == 'POST':
        form = TaskForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('home')
    else:
        form = TaskForm()

    # 1. Empezamos con todas las tareas
    tasks = models.Task.objects.all().order_by('-created_at')

    # 2. Capturamos los filtros desde la URL (GET)
    f_fecha = request.GET.get('fecha')
    f_proceso = request.GET.get('proceso')
    f_estado = request.GET.get('estado')

    # 3. Aplicamos los filtros de forma encadenada si existen
    if f_fecha:
        tasks = tasks.filter(created_at__date=f_fecha)
    
    if f_proceso:
        # 'icontains' para que busque parte del nombre sin importar mayúsculas
        tasks = tasks.filter(proceso=f_proceso)
    
    if f_estado:
        tasks = tasks.filter(state=f_estado)

    return render(request, 'index.html', {
        'tasks': tasks,
        'form': form,
        'lista_procesos': lista_procesos,
    })

def completar_tarea(request, task_id):
    task = get_object_or_404(models.Task, id=task_id)
    task.state = 'C' 
    task.save()
    return redirect('home')

def agregar_solucion(request, task_id):
    if request.method == 'POST':
        tarea = get_object_or_404(models.Task, id=task_id)
        tarea.solucion = request.POST.get('solucion_texto')
        tarea.save()
    return redirect('home') 

def exportar_tareas_csv(request):
    # 1. Aplicamos la misma lógica de filtrado que en tu vista 'home'
    tasks = models.Task.objects.all().order_by('-created_at')

    f_fecha = request.GET.get('fecha')
    f_proceso = request.GET.get('proceso')
    f_estado = request.GET.get('estado')

    if f_fecha:
        tasks = tasks.filter(created_at__date=f_fecha)
    if f_proceso:
        tasks = tasks.filter(proceso_id=f_proceso)
    if f_estado:
        tasks = tasks.filter(state=f_estado)

    # 2. Creamos el objeto de respuesta CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="reporte_tareas_{f_fecha or "general"}.csv"'
    
    # Manejo de caracteres especiales (acentos)
    response.write(u'\ufeff'.encode('utf8'))
    writer = csv.writer(response)
    
    # Cabeceras
    writer.writerow(['ID', 'Fecha', 'Proceso', 'Reportado Por', 'Prioridad', 'Estado', 'Descripción', 'Solución'])

    # Datos
    for t in tasks:
        writer.writerow([
            t.id,
            t.created_at.strftime('%d/%m/%Y %H:%M'),
            t.proceso.nombre,
            t.reportado_por,
            t.get_priority_display(),
            t.get_state_display(),
            t.descripcion,
            t.solucion if t.solucion else "Sin solución"
        ])

    return response


def exportar_tareas_pdf(request):
    tasks = models.Task.objects.all().order_by('-created_at')
    
    f_fecha = request.GET.get('fecha')
    f_proceso = request.GET.get('proceso')
    f_estado = request.GET.get('estado')

    if f_fecha:
        tasks = tasks.filter(created_at__date=f_fecha)
    if f_proceso:
        tasks = tasks.filter(proceso_id=f_proceso)
    if f_estado:
        tasks = tasks.filter(state=f_estado)

    template_path = 'reporte_pdf.html'
    context = {
        'tasks': tasks,
        'hoy': timezone.now(),
    }
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_{f_fecha or "general"}.pdf"'
    
    template = get_template(template_path)
    html = template.render(context)

    pisa_status = pisa.CreatePDF(html, dest=response)
    
    if pisa_status.err:
       return HttpResponse('Ocurrió un error al generar el PDF', status=500)
    return response

def exportar_tareas_excel(request):
    tasks = models.Task.objects.all().order_by('-created_at')
    
    f_fecha = request.GET.get('fecha')
    f_proceso = request.GET.get('proceso')
    f_estado = request.GET.get('estado')

    if f_fecha:
        tasks = tasks.filter(created_at__date=f_fecha)
    if f_proceso:
        tasks = tasks.filter(proceso_id=f_proceso)
    if f_estado:
        tasks = tasks.filter(state=f_estado)

    wb = Workbook()
    ws = wb.active
    ws.title = "Registro de Solicitudes"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Azul oscuro
    header_font = Font(bold=True, color="FFFFFF", size=12)
    center_ali = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ['ID', 'FECHA', 'PROCESO', 'REPORTADO POR', 'PRIORIDAD', 'ESTADO', 'DESCRIPCIÓN', 'SOLUCIÓN']
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_ali
        cell.border = border

    for t in tasks:
        row = [
            t.id,
            t.created_at.strftime('%d/%m/%Y %H:%M'),
            t.proceso.nombre,
            t.reportado_por,
            t.get_priority_display(),
            t.get_state_display(),
            t.descripcion,
            t.solucion if t.solucion else "Pendiente"
        ]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True) 

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 50)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Registro_Solicitudes_{f_fecha or "General"}.xlsx"'
    
    wb.save(response)
    return response